


from datetime import datetime
start_time = datetime.now()

###Code




###### define mod value
mod=5000



from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")



##### imorting openpyxl and nan and loading workbook
try:
    from cmath import nan
    import openpyxl

    wb = openpyxl.load_workbook(r'C:\Users\DELL\OneDrive\Documents\GitHub\2001ME67_2022\tut02\input_octant_transition_identify.xlsx')
except:
    print("there is error in loading workbook check your file directory and import openpyxl")
    exit()


    #### calculating average value
try:
    sheet = wb.active
    Uavg=0
    Vavg=0
    Wavg=0
    ls=[]
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=29746, max_col=4):
        lst1=[]
        for cell in row:
            lst1.append(cell.value)
        ls.append(lst1)
    for i in range(1,29746):
        Uavg=Uavg+ls[i][1]
        Vavg=Vavg+ls[i][2]
        Wavg=Wavg+ls[i][3]
    Uavg=Uavg/29745
    Vavg=Vavg/29745
    Wavg=Wavg/29745

    lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
except:
    print("there is error in calculating average value")
    exit()

### writing nan in tansition value

try:
    i=0
    for row in sheet.iter_rows(min_row=1, min_col=12, max_row=29777, max_col=21):
        j=0
        for cell in row:
            cell.value=nan
            j=j+1
        i=i+1
except:
    print("there is some error in transition count")
#####creating octant of given value

try:
    i=0
    for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
        j=0
        for cell in row:
            cell.value=lst_avg[i][j]
            j=j+1
        i=i+1
    lst_newval=[]
    for i in range(1,29746):
        lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])
    sheet.cell(row=1,column=8).value="Uavg'"
    sheet.cell(row=1,column=9).value="Vavg'"
    sheet.cell(row=1,column=10).value="Wavg'"    
    sheet.cell(row=1,column=11).value="Octant"    

    i=0
    for row in sheet.iter_rows(min_row=2, min_col=8, max_row=29746, max_col=10):
        j=0
        for cell in row:
            cell.value=lst_newval[i][j]
            j=j+1
        i=i+1

    lst_octant = []
    for p in lst_newval:
        if(p[0]>=0):
            if(p[1]>=0):
                if(p[2]>=0):
                    lst_octant.append(1)
                else:
                    lst_octant.append(-1)
            else:
                if(p[2]>=0):
                    lst_octant.append(4)
                else:
                    lst_octant.append(-4)
        else:
            if(p[1]>=0):
                if(p[2]>=0):
                    lst_octant.append(2)
                else:
                    lst_octant.append(-2)
            else:
                if(p[2]>=0):
                    lst_octant.append(3)
                else:
                    lst_octant.append(-3)

    i=0
    for row in sheet.iter_rows(min_row=2, min_col=11, max_row=29746, max_col=11):
        for cell in row:
            cell.value=lst_octant[i]
        i=i+1
except:
    print("there is error in creating octant value")
    exit()

###### there is some error in creating overall count

try:
    t=0
    if(29746%mod==0):
        t=29746//mod
    else:
        t=(29746//mod)+1

    sheet.cell(row=1,column=14).value="+1"
    sheet.cell(row=1,column=15).value="-1"
    sheet.cell(row=1,column=16).value="+2"
    sheet.cell(row=1,column=17).value="-2"
    sheet.cell(row=1,column=18).value="+3"
    sheet.cell(row=1,column=19).value="-3"
    sheet.cell(row=1,column=20).value="+4"
    sheet.cell(row=1,column=21).value="-4"
    sheet.cell(row=2,column=13).value="Overall Count"
    sheet.cell(row=3,column=12).value="User Input"
    tt=str(mod)
    sheet.cell(row=3,column=13).value="mod"+" "+tt
    lst_overall_count = [0,0,0,0,0,0,0,0]
    for valu in lst_octant:
        if valu==1:
            lst_overall_count[0]=lst_overall_count[0]+1
        if valu==-1:
            lst_overall_count[1]=lst_overall_count[1]+1
        if valu==2:
            lst_overall_count[2]=lst_overall_count[2]+1
        if valu==-2:
            lst_overall_count[3]=lst_overall_count[3]+1
        if valu==3:
            lst_overall_count[4]=lst_overall_count[4]+1
        if valu==-3:
            lst_overall_count[5]=lst_overall_count[5]+1
        if valu==4:
            lst_overall_count[6]=lst_overall_count[6]+1
        if valu==-4:
            lst_overall_count[7]=lst_overall_count[7]+1
    for row in sheet.iter_rows(min_row=2, min_col=14, max_row=2, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_overall_count[j]
            j=j+1
    lst_hh=[]
    for j in range(t):
        if(j==t-1):
            ttm1=str(mod*j)
            ttm=ttm1+"-"+"29745"
            lst_hh.append(ttm)
        else:
            if(j==0):
                ttm1=".0000"
            else:
                ttm1=str(mod*j) 
            ttm2=str(mod*(j+1)-1)
            ttm=ttm1+"-"+ttm2
            lst_hh.append(ttm) 
    lst_hh.append("Verified")
except:
    print("there is error in creating overall count value and verifing it")
    exit()

##### updating overall count and transition
try:
    lst_hh_val=[]
    for j in range(t):
        lst_hh_temp=[0,0,0,0,0,0,0,0]
        if(j==t-1):
            y=29745
        else:
            y=mod*(j+1)
        for valu in range(mod*j,y):
            if lst_octant[valu]==1:
                lst_hh_temp[0]=lst_hh_temp[0]+1
            if lst_octant[valu]==-1:
                lst_hh_temp[1]=lst_hh_temp[1]+1
            if lst_octant[valu]==2:
                lst_hh_temp[2]=lst_hh_temp[2]+1
            if lst_octant[valu]==-2:
                lst_hh_temp[3]=lst_hh_temp[3]+1
            if lst_octant[valu]==3:
                lst_hh_temp[4]=lst_hh_temp[4]+1
            if lst_octant[valu]==-3:
                lst_hh_temp[5]=lst_hh_temp[5]+1
            if lst_octant[valu]==4:
                lst_hh_temp[6]=lst_hh_temp[6]+1
            if lst_octant[valu]==-4:
                lst_hh_temp[7]=lst_hh_temp[7]+1


        lst_hh_val.append(lst_hh_temp)

    lst_verified=[0,0,0,0,0,0,0,0]
    for i in range(t):
        for j in range(8):
            lst_verified[j]=lst_verified[j]+lst_hh_val[i][j]

    j=0
    for row in sheet.iter_rows(min_row=4, min_col=13, max_row=t+4, max_col=13):
        for cell in row:
            cell.value=lst_hh[j]
        j=j+1

    i=0
    for row in sheet.iter_rows(min_row=4, min_col=14, max_row=t+3, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_hh_val[i][j]
            j=j+1
        i=i+1

    for row in sheet.iter_rows(min_row=t+4, min_col=14, max_row=t+4, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_verified[j]
            j=j+1
except:
    print("there is error in updating in verified count")


#### creating list for mod transition value
try:
    lst_tran_count=[]

    for i in range(t):
        if(i==t-1):
            y=29744
        else:
            y=mod*(i+1)-1
        lst_tran_count_temp=[]
        temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        for k in range(mod*i,y):

            if(lst_octant[k]==1):
                temp_1[lst_octant[k+1]]=temp_1[lst_octant[k+1]]+1
            if(lst_octant[k]==-1):
                temp_11[lst_octant[k+1]]=temp_11[lst_octant[k+1]]+1
            if(lst_octant[k]==2):
                temp_2[lst_octant[k+1]]=temp_2[lst_octant[k+1]]+1
            if(lst_octant[k]==-2):
                temp_22[lst_octant[k+1]]=temp_22[lst_octant[k+1]]+1
            if(lst_octant[k]==3):
                temp_3[lst_octant[k+1]]=temp_3[lst_octant[k+1]]+1
            if(lst_octant[k]==-3):
                temp_33[lst_octant[k+1]]=temp_33[lst_octant[k+1]]+1
            if(lst_octant[k]==4):
                temp_4[lst_octant[k+1]]=temp_4[lst_octant[k+1]]+1
            if(lst_octant[k]==-4):
                temp_44[lst_octant[k+1]]=temp_44[lst_octant[k+1]]+1
        lst_tran_count_temp.append(temp_1)
        lst_tran_count_temp.append(temp_11)
        lst_tran_count_temp.append(temp_2)
        lst_tran_count_temp.append(temp_22)
        lst_tran_count_temp.append(temp_3)
        lst_tran_count_temp.append(temp_33)
        lst_tran_count_temp.append(temp_4)
        lst_tran_count_temp.append(temp_44)
        lst_tran_count.append(lst_tran_count_temp)
        lst_tran_count_temp=[]
        temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}

    lst_tran_count1=[]

    for i in range(t):

        lst_tran_count_temp=[]
        temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        for k in range(29744):

            if(lst_octant[k]==1):
                temp_1[lst_octant[k+1]]=temp_1[lst_octant[k+1]]+1
            if(lst_octant[k]==-1):
                temp_11[lst_octant[k+1]]=temp_11[lst_octant[k+1]]+1
            if(lst_octant[k]==2):
                temp_2[lst_octant[k+1]]=temp_2[lst_octant[k+1]]+1
            if(lst_octant[k]==-2):
                temp_22[lst_octant[k+1]]=temp_22[lst_octant[k+1]]+1
            if(lst_octant[k]==3):
                temp_3[lst_octant[k+1]]=temp_3[lst_octant[k+1]]+1
            if(lst_octant[k]==-3):
                temp_33[lst_octant[k+1]]=temp_33[lst_octant[k+1]]+1
            if(lst_octant[k]==4):
                temp_4[lst_octant[k+1]]=temp_4[lst_octant[k+1]]+1
            if(lst_octant[k]==-4):
                temp_44[lst_octant[k+1]]=temp_44[lst_octant[k+1]]+1
        lst_tran_count_temp.append(temp_1)
        lst_tran_count_temp.append(temp_11)
        lst_tran_count_temp.append(temp_2)
        lst_tran_count_temp.append(temp_22)
        lst_tran_count_temp.append(temp_3)
        lst_tran_count_temp.append(temp_33)
        lst_tran_count_temp.append(temp_4)
        lst_tran_count_temp.append(temp_44)
        lst_tran_count1.append(lst_tran_count_temp)
        lst_tran_count_temp=[]
        temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
except:
    print("there is some error in creating list of mod transition value")


#### updating mod transition value
try:
    lst_skelton_1=[[nan,"Overall Transition Count",nan,nan,nan,nan,nan,nan,nan,nan],[nan,nan,"To",nan,nan,nan,nan,nan,nan,nan],[nan,"count","+1","-1","+2","-2","+3","-3","+4","-4"],["From","+1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+4",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-4",nan,nan,nan,nan,nan,nan,nan,nan]]
    lst_skelton=[[nan,"Mod Transition count",nan,nan,nan,nan,nan,nan,nan,nan],[nan,nan,"To",nan,nan,nan,nan,nan,nan,nan],[nan,"count","+1","-1","+2","-2","+3","-3","+4","-4"],["From","+1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+4",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-4",nan,nan,nan,nan,nan,nan,nan,nan]]
    i=0
    zz=t+7
    for row in sheet.iter_rows(min_row=zz, min_col=12, max_row=zz+10, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_skelton_1[i][j]
            j=j+1
        i=i+1
    for h in range(1,t+1):
        i=0
        for row in sheet.iter_rows(min_row=13*h+zz, min_col=12, max_row=13*(h)+zz+10, max_col=21):
            j=0
            for cell in row:
                cell.value=lst_skelton[i][j]
                j=j+1
            i=i+1

    lst_tran_count_ans=[]
    for lsss in lst_tran_count:
        list_temp_1=[]
        for lsss1 in lsss:
            list_temp=[]
            for x in lsss1:
                list_temp.append(lsss1[x])
            list_temp_1.append(list_temp)
        lst_tran_count_ans.append(list_temp_1)
    lst_tran_count_ans1=[]
    for lsss in lst_tran_count1:
        list_temp_1=[]
        for lsss1 in lsss:
            list_temp=[]
            for x in lsss1:
                list_temp.append(lsss1[x])
            list_temp_1.append(list_temp)
        lst_tran_count_ans1.append(list_temp_1)


    for h in range(1,t+1):
        i=0
        for row in sheet.iter_rows(min_row=13*(h)+zz+3, min_col=14, max_row=13*(h)+zz+10, max_col=21):
            j=0
            for cell in row:
                cell.value=lst_tran_count_ans[h-1][i][j]
                j=j+1
            i=i+1
    i=0
    for h in range(1,t+1):
        for row in sheet.iter_rows(min_row=13*(h)+zz+1, min_col=13, max_row=13*(h)+zz+1, max_col=13):
            for cell in row:
                cell.value=lst_hh[i]    
            i=i+1

    i=0
    for row in sheet.iter_rows(min_row=zz+3, min_col=14, max_row=zz+10, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_tran_count_ans1[0][i][j]
            j=j+1
        i=i+1
except:
    print("there is error in updating mod transition value in excel")



##### saving workbook file
try:
    wb.save(r'C:\Users\DELL\OneDrive\Documents\GitHub\2001ME67_2022\tut02\output_octant_transition_identify.xlsx')
except:
    print("there is error in saving excel file")

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))

