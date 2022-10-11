
from datetime import datetime
start_time = datetime.now()

from platform import python_version
ver = python_version()



###### define mod value
mod=5000




if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

######code
def octant_range_names(mod=5000):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}



    ##### imorting openpyxl and nan and loading workbook
    try:
        from cmath import nan
        import openpyxl

        wb = openpyxl.load_workbook(r'C:\Users\DELL\OneDrive\Desktop\octant ID\octant_input.xlsx')
    except:
        print("there is error in loading workbook check your file directory and import openpyxl")
        exit()


    
    #### calculating average value
    try:
        sheet = wb.active
        sheet.column_dimensions['O'].width=40
        sheet.column_dimensions['AE'].width=40
        Uavg=0
        Vavg=0
        Wavg=0
        ls=[]
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=29746, max_col=4):
            lst1=[]
            for cell in row:
                lst1.append(cell.value)
            ls.append(lst1)
        for i in range(1,29746):
            Uavg=Uavg+ls[i][1]
            Vavg=Vavg+ls[i][2]
            Wavg=Wavg+ls[i][3]
        Uavg=Uavg/29745
        Vavg=Vavg/29745
        Wavg=Wavg/29745

        lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
    except:
        print("there is error in calculating average value")
        exit()

    ### writing nan in tansition value

    try:
        i=0
        for row in sheet.iter_rows(min_row=1, min_col=12, max_row=29777, max_col=51):
            j=0
            for cell in row:
                cell.value=nan
                j=j+1
            i=i+1
    except:
        print("there is some error in transition count")
    

    #####creating octant of given value

    try:
        i=0
        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
            j=0
            for cell in row:
                cell.value=lst_avg[i][j]
                j=j+1
            i=i+1
        lst_newval=[]
        for i in range(1,29746):
            lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])
        sheet.cell(row=1,column=8).value="Uavg'"
        sheet.cell(row=1,column=9).value="Vavg'"
        sheet.cell(row=1,column=10).value="Wavg'"    
        sheet.cell(row=1,column=11).value="Octant"    

        i=0
        for row in sheet.iter_rows(min_row=2, min_col=8, max_row=29746, max_col=10):
            j=0
            for cell in row:
                cell.value=lst_newval[i][j]
                j=j+1
            i=i+1

        lst_octant = []
        for p in lst_newval:
            if(p[0]>=0):
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(1)
                    else:
                        lst_octant.append(-1)
                else:
                    if(p[2]>=0):
                        lst_octant.append(4)
                    else:
                        lst_octant.append(-4)
            else:
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(2)
                    else:
                        lst_octant.append(-2)
                else:
                    if(p[2]>=0):
                        lst_octant.append(3)
                    else:
                        lst_octant.append(-3)

        i=0
        for row in sheet.iter_rows(min_row=2, min_col=11, max_row=29746, max_col=11):
            for cell in row:
                cell.value=lst_octant[i]
            i=i+1
    except:
        print("there is error in creating octant value")
        exit()
    
    ###### creating overall count

    try:
        t=0
        if(29746%mod==0):
            t=29746//mod
        else:
            t=(29746//mod)+1

        sheet.cell(row=2,column=14).value=1
        sheet.cell(row=2,column=15).value=-1
        sheet.cell(row=2,column=16).value=2
        sheet.cell(row=2,column=17).value=-2
        sheet.cell(row=2,column=18).value=3
        sheet.cell(row=2,column=19).value=-3
        sheet.cell(row=2,column=20).value=4
        sheet.cell(row=2,column=21).value=-4
        sheet.cell(row=3,column=13).value="Overall Count"
        sheet.cell(row=4,column=12).value="User Input"
        tt=str(mod)
        sheet.cell(row=4,column=13).value="mod"+" "+tt
        lst_overall_count = [0,0,0,0,0,0,0,0]
        for valu in lst_octant:
            if valu==1:
                lst_overall_count[0]=lst_overall_count[0]+1
            if valu==-1:
                lst_overall_count[1]=lst_overall_count[1]+1
            if valu==2:
                lst_overall_count[2]=lst_overall_count[2]+1
            if valu==-2:
                lst_overall_count[3]=lst_overall_count[3]+1
            if valu==3:
                lst_overall_count[4]=lst_overall_count[4]+1
            if valu==-3:
                lst_overall_count[5]=lst_overall_count[5]+1
            if valu==4:
                lst_overall_count[6]=lst_overall_count[6]+1
            if valu==-4:
                lst_overall_count[7]=lst_overall_count[7]+1
        for row in sheet.iter_rows(min_row=3, min_col=14, max_row=3, max_col=21):
            j=0
            for cell in row:
                cell.value=lst_overall_count[j]
                j=j+1
        lst_hh=[]
        for j in range(t):
            if(j==t-1):
                ttm1=str(mod*j)
                ttm=ttm1+"-"+"29745"
                lst_hh.append(ttm)
            else:
                if(j==0):
                    ttm1=".0000"
                else:
                    ttm1=str(mod*j) 
                ttm2=str(mod*(j+1)-1)
                ttm=ttm1+"-"+ttm2
                lst_hh.append(ttm) 

    except:
        print("there is error in creating overall count value")
        exit()

    ##### updating overall count and transition
    try:
        lst_hh_val=[]
        for j in range(t):
            lst_hh_temp=[0,0,0,0,0,0,0,0]
            if(j==t-1):
                y=29745
            else:
                y=mod*(j+1)
            for valu in range(mod*j,y):
                if lst_octant[valu]==1:
                    lst_hh_temp[0]=lst_hh_temp[0]+1
                if lst_octant[valu]==-1:
                    lst_hh_temp[1]=lst_hh_temp[1]+1
                if lst_octant[valu]==2:
                    lst_hh_temp[2]=lst_hh_temp[2]+1
                if lst_octant[valu]==-2:
                    lst_hh_temp[3]=lst_hh_temp[3]+1
                if lst_octant[valu]==3:
                    lst_hh_temp[4]=lst_hh_temp[4]+1
                if lst_octant[valu]==-3:
                    lst_hh_temp[5]=lst_hh_temp[5]+1
                if lst_octant[valu]==4:
                    lst_hh_temp[6]=lst_hh_temp[6]+1
                if lst_octant[valu]==-4:
                    lst_hh_temp[7]=lst_hh_temp[7]+1


            lst_hh_val.append(lst_hh_temp)
        for i in range(t):
            sheet.cell(row=i+5,column=13).value=lst_hh[i]
        for i in range(t):
            for j in range(8):
                sheet.cell(row=i+5,column=14+j).value=lst_hh_val[i][j]

    except:
        print("there is error in updating in verified count")
        
    #### creating skelton and list of rank
    try:
        lst_rec=[1,-1,2,-2,3,-3,4,-4]
        lst_rec1=["1","-1","2","-2","3","-3","4","-4"]
        lst_rank_head=["Rank 1","Rank 2","Rank 3","Rank 4","Rank 5","Rank 6","Rank 7","Rank 8","Rank1 OctantID","Rank1 Octant Name"]
        for i in range(len(lst_rec)):
            sheet.cell(row=1,column=i+22).value=lst_rec[i]
        for i in range(len(lst_rank_head)):
            sheet.cell(row=2,column=i+22).value=lst_rank_head[i]
        lst_record=[]
        lst_overall_count1=lst_overall_count.copy()
        lst_overall_count1.sort(reverse=True)
        list_temp=[]
        for i in range(8):
            for j in range(8):
                if(lst_overall_count[i]==lst_overall_count1[j]):
                    if(j==0):
                        lst_record.append(i)
                    list_temp.append(j+1)
                    break
        for i in range(8):
            sheet.cell(row=3,column=i+22).value=list_temp[i]
        sheet.cell(row=3,column=30).value=lst_record[0]
        t1=str(lst_rec[lst_record[0]])
        sheet.cell(row=3,column=31).value=octant_name_id_mapping[t1]

        lst_rank=[]
        lst_record=[]
        for i in range(t):
            lst_hh_val1=lst_hh_val[i].copy()
            lst_hh_val1.sort(reverse=True)
            list_temp=[]
            for k in range(8):
                for j in range(8):
                    if(lst_hh_val[i][k]==lst_hh_val1[j]):
                        if(j==0):
                            lst_record.append(k)
                        list_temp.append(j+1)
                        break
            lst_rank.append(list_temp)
        for i in range(t):
            for j in range(8):
                sheet.cell(row=i+5,column=j+22).value=lst_rank[i][j]
        lst_record1=[]
        lst_record2=[]
        for item in lst_record:
            lst_record1.append(str(lst_rec[item]))
            lst_record2.append((lst_rec[item]))
        for i in range(len(lst_record2)):

        
    except:
        print("there is some errror in creating skelton and list of rank")
    ##### saving workbook file
    try:
        wb.save(r'C:\Users\DELL\OneDrive\Desktop\octant ID\octant_output_ranking_excel1.xlsx')
    except:
        print("there is error in saving excel file")
    

    
   
octant_range_names(mod)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))