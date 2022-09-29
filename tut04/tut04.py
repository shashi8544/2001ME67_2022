#### importing openpyxl and nan


try:
    from cmath import nan
    import openpyxl

    wb = openpyxl.load_workbook(r'C:\Users\DELL\OneDrive\Desktop\octant_longest_subsequene_tut3\input_octant_longest_subsequence_with_range.xlsx')
    sheet = wb.active
    sheet.column_dimensions['N'].width=35
    sheet.column_dimensions['R'].width=35
except:
    print("there is error in loading workbook check your file directory and import openpyxl")
    exit()


    #### calculating average value

try:
    Uavg=0
    Vavg=0
    Wavg=0
    ls=[]
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=29746, max_col=4):
        lst1=[]
        for cell in row:
            lst1.append(cell.value)
        ls.append(lst1)
    for i in range(1,29746):
        Uavg=Uavg+ls[i][1]
        Vavg=Vavg+ls[i][2]
        Wavg=Wavg+ls[i][3]
    Uavg=Uavg/29745
    Vavg=Vavg/29745
    Wavg=Wavg/29745

    lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
except:
    print("there is error in calculating average value")
    exit()


i=0
for row in sheet.iter_rows(min_row=2, min_col=8, max_row=29746, max_col=10):
    j=0
    for cell in row:
        cell.value=nan
        j=j+1
    i=i+1
#####creating octant of given value

try:
    i=0
    for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
        j=0
        for cell in row:
            cell.value=lst_avg[i][j]
            j=j+1
        i=i+1
    lst_newval=[]
    for i in range(1,29746):
        lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])
    sheet.cell(row=1,column=8).value="Uavg'"
    sheet.cell(row=1,column=9).value="Vavg'"
    sheet.cell(row=1,column=10).value="Wavg'"    
    sheet.cell(row=1,column=11).value="Octant"    

    i=0
    for row in sheet.iter_rows(min_row=2, min_col=8, max_row=29746, max_col=10):
        j=0
        for cell in row:
            cell.value=lst_newval[i][j]
            j=j+1
        i=i+1

    lst_octant = []
    for p in lst_newval:
        if(p[0]>=0):
            if(p[1]>=0):
                if(p[2]>=0):
                    lst_octant.append(1)
                else:
                    lst_octant.append(-1)
            else:
                if(p[2]>=0):
                    lst_octant.append(4)
                else:
                    lst_octant.append(-4)
        else:
            if(p[1]>=0):
                if(p[2]>=0):
                    lst_octant.append(2)
                else:
                    lst_octant.append(-2)
            else:
                if(p[2]>=0):
                    lst_octant.append(3)
                else:
                    lst_octant.append(-3)

    i=0
    for row in sheet.iter_rows(min_row=2, min_col=11, max_row=29746, max_col=11):
        for cell in row:
            cell.value=lst_octant[i]
        i=i+1
except:
    print("there is error in creating octant value")
    exit()



### creating skelton table
try:
    row_head = ["count","Longest Subsquence Length","count"]
    row_column = ["+1","-1","+2","-2","+3","-3","+4","-4"]
    for row in sheet.iter_rows(min_row=2, min_col=13, max_row=2, max_col=15):
        j=0
        for cell in row:
            cell.value=row_head[j]
            j=j+1
    for row in sheet.iter_rows(min_row=2, min_col=17, max_row=2, max_col=19):
        j=0
        for cell in row:
            cell.value=row_head[j]
            j=j+1

    i=0
    for row in sheet.iter_rows(min_row=3, min_col=13, max_row=10, max_col=13):
        for cell in row:
            cell.value=row_column[i]
        i=i+1
except:
    print("there is error in creating skelton of table")




#####  subsequence codes 
try:
    prev1=[1,-1,2,-2,3,-3,4,-4]
    ans=[]

    for prev in prev1:
        count1_max=0
        count1=0
        temp_count=0
        for v in lst_octant:
            if(v==prev):
                temp_count=temp_count+1
            else:
                if(temp_count>count1_max):
                    count1=1
                    count1_max=temp_count
                elif(count1_max==temp_count & count1_max!=0):
                    count1=count1+1
                temp_count=0
        lst_temp=[count1_max,count1]
        ans.append(lst_temp)
        count1_max=0
        count1=0
        temp_count=0
except:
    print("there is some error in codes of subsequence")


try:
    wb.save(r'C:\Users\DELL\OneDrive\Desktop\octant_longest_subsequene_tut3\input_octant_longest_subsequence_with_range.xlsx')
except:
    print("there is some error in saving excel file")